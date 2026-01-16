import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from pathlib import Path
import sys
import csv
import re
import html
import unicodedata
from datetime import datetime

csv.field_size_limit(50 * 1024 * 1024)  # 50 MB

# =====================================================
# ================= COUNT RECORDS =====================
# =====================================================
def count_bibtex_records(bib_path):
    count = 0
    with open(bib_path, "r", encoding="utf-8") as f:
        for line in f:
            if line.strip().startswith("@"):
                count += 1
    return count

def count_csv_records(csv_path):
    with open(csv_path, "r", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        rows = list(reader)
        return len(rows) - 1 if len(rows) > 0 else 0

# =====================================================
# ================= PATHS =============================
# =====================================================
zenodo_csv = r"C:\Users\maria\Desktop\Cloud-Ontology\Fetcher-Functions\output-zenodo\zenodo_all_years.csv"
output_dir = Path(r"C:\Users\maria\Desktop\Cloud-Ontology\Fetcher-Functions\output-zenodo")

zenodo_xlsx = output_dir / "zenodo_all_years.xlsx"
duplicates_csv = output_dir / "zenodo_duplicates_removed.csv"
bibtex_path = output_dir / "zenodo_all_years.bib"

# =====================================================
# ================= CSV UTILS =========================
# =====================================================
def detect_separator(file_path):
    with open(file_path, 'r', encoding='utf-8') as f:
        sample = f.read(4096)
        try:
            return csv.Sniffer().sniff(sample, delimiters=[',', ';', '\t']).delimiter
        except Exception:
            return ','

def read_csv_stable(file_path):
    sep = detect_separator(file_path)
    df = pd.read_csv(file_path, encoding='utf-8', sep=sep, on_bad_lines='skip')
    print(f"[INFO] CSV letto: {len(df)} righe")
    return df

# =====================================================
# ================= TEXT CLEANING =====================
# =====================================================
def clean_text(s):
    """
    Pulizia HTML aggressiva:
    - rimuove script e style
    - rimuove tutti i tag HTML
    - decodifica ed elimina entità HTML (&nbsp;, &amp;, &#160;, ecc.)
    - normalizza Unicode
    - rimuove spazi, newline e caratteri invisibili
    """
    if pd.isna(s):
        return ""

    s = str(s)

    # 1. Rimuove script e style
    s = re.sub(
        r'<(script|style).*?>.*?</\1>',
        '',
        s,
        flags=re.DOTALL | re.IGNORECASE
    )

    # 2. Rimuove tutti i tag HTML
    s = re.sub(r'<[^>]+>', ' ', s)

    # 3. Decodifica entità HTML
    s = html.unescape(s)

    # 4. Rimuove entità residue
    s = re.sub(r'&[a-zA-Z0-9#]+;', ' ', s)

    # 5. Normalizzazione Unicode
    s = unicodedata.normalize("NFKC", s)

    # 6. Rimuove caratteri invisibili
    s = re.sub(r'[\x00-\x1F\x7F]', ' ', s)

    # 7. Collassa spazi
    s = re.sub(r'\s+', ' ', s)

    return s.strip()

def normalize_title(title):
    title = (title or "").lower()
    title = re.sub(r"[^\w\s]", "", title)
    return re.sub(r"\s+", " ", title).strip()

def parse_date(d):
    try:
        return datetime.fromisoformat(str(d).replace("Z", ""))
    except Exception:
        return datetime.min

# =====================================================
# ================= DEDUPLICATION =====================
# =====================================================
def deduplicate(df):
    unique = {}
    duplicates = []

    for _, row in df.iterrows():
        title = normalize_title(row.get("title"))
        authors = normalize_title(row.get("authors"))
        doi = str(row.get("doi", "")).strip().lower()
        url = str(row.get("url", "")).strip().lower()

        if title and authors:
            key = f"title_auth::{title}::{authors}"
        elif doi:
            key = f"doi::{doi}"
        elif url:
            key = f"url::{url}"
        else:
            continue

        if key not in unique:
            unique[key] = row
        else:
            duplicates.append(row)

    return pd.DataFrame(unique.values()), pd.DataFrame(duplicates)

# =====================================================
# ================= EXCEL FORMAT ======================
# =====================================================
def format_excel_table(excel_path):
    wb = load_workbook(excel_path)
    ws = wb.active

    max_row = ws.max_row
    max_col = ws.max_column
    last_col = get_column_letter(max_col)

    table = Table(
        displayName="ZenodoTable",
        ref=f"A1:{last_col}{max_row}"
    )
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showRowStripes=True
    )
    ws.add_table(table)

    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border

    for col in ws.columns:
        letter = col[0].column_letter
        max_len = 0
        for cell in col:
            if cell.row == 1 or not cell.value:
                continue
            val = str(cell.value)
            max_len = max(max_len, len(val))
            if val.startswith("10."):
                cell.hyperlink = f"https://doi.org/{val}"
                cell.font = Font(color="0000EE", underline="single")
            elif val.startswith("http"):
                cell.hyperlink = val
                cell.font = Font(color="0000EE", underline="single")
        ws.column_dimensions[letter].width = min(max_len + 2, 60)

    wb.save(excel_path)

# =====================================================
# ================= BIBTEX EXPORT =====================
# =====================================================
def export_bibtex(df, path):
    with open(path, "w", encoding="utf-8") as f:
        for _, r in df.iterrows():
            key = r.get("doi") or r.get("url") or normalize_title(r.get("title"))[:40]
            year = str(r.get("created", ""))[:4]

            f.write(
                f"@misc{{{key},\n"
                f"  title = {{{r.get('title')}}},\n"
                f"  author = {{{r.get('authors')}}},\n"
                f"  year = {{{year}}},\n"
                f"  howpublished = {{Zenodo}},\n"
                f"  url = {{{r.get('url')}}}\n"
                f"}}\n\n"
            )

# =====================================================
# ================= PIPELINE ==========================
# =====================================================
print("\n📚 Post-processing Zenodo finale")

df = read_csv_stable(zenodo_csv)

for col in df.columns:
    df[col] = df[col].apply(clean_text)

df_clean, df_duplicates = deduplicate(df)

df_clean.to_excel(zenodo_xlsx, index=False)
df_duplicates.to_csv(duplicates_csv, index=False, encoding="utf-8-sig")

format_excel_table(zenodo_xlsx)
export_bibtex(df_clean, bibtex_path)

print("\n✅ COMPLETATO")
print(f"✔ Record finali: {len(df_clean)}")
print(f"🗑️ Duplicati rimossi: {len(df_duplicates)}")
print(f"📄 Excel: {zenodo_xlsx}")
print(f"📚 BibTeX: {bibtex_path}")

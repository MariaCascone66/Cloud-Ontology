import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from pathlib import Path
import csv
import re

# === Percorsi ===
files = {
    r"C:\Users\maria\Desktop\Cloud-Ontology\Fetcher-Results\github_combined.csv":
        r"C:\Users\maria\Desktop\Replication package\File Excel\output_github.xlsx",
    r"C:\Users\maria\Desktop\Cloud-Ontology\Fetcher-Results\lodcloud_results.csv":
        r"C:\Users\maria\Desktop\Replication package\File Excel\output_lodcloud.xlsx",
    r"C:\Users\maria\Desktop\Replication package\Biblioteca\Biblioteca.csv":
        r"C:\Users\maria\Desktop\Replication package\File Excel\output_zotero.xlsx",
}

zenodo_csv = r"C:\Users\maria\Desktop\Cloud-Ontology\Fetcher-Results\zenodo_combined.csv"
zenodo_xlsx = r"C:\Users\maria\Desktop\Replication package\File Excel\output_zenodo.xlsx"


# === Utility CSV ===
def detect_separator(file_path):
    with open(file_path, 'r', encoding='utf-8') as f:
        sample = f.read(4096)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=[',', ';', '\t'])
            return dialect.delimiter
        except Exception:
            return ','


def read_csv_stable(file_path):
    sep = detect_separator(file_path)
    print(f"[INFO] Rilevato separatore '{sep}' per {file_path}")
    df = pd.read_csv(file_path, encoding='utf-8', sep=sep, on_bad_lines='skip')
    print(f"[INFO] Letto CSV: {len(df)} righe, {len(df.columns)} colonne")
    return df


# === Funzione per pulire testo (HTML, spazi, newline) ===
def clean_text(s):
    if pd.isna(s):
        return ''
    s = re.sub(r'<[^>]+>', '', str(s))
    s = re.sub(r'\s+', ' ', s).strip()
    return s


# === Funzione di formattazione Excel e hyperlink ===
def format_excel_table(excel_path):
    wb = load_workbook(excel_path)
    ws = wb.active

    max_row = ws.max_row
    max_col = ws.max_column
    last_col = get_column_letter(max_col)
    table_range = f"A1:{last_col}{max_row}"

    # Tabella con stile
    table = Table(displayName="DataTable", ref=table_range)
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)

    # Bordo e allineamento
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border

    # Hyperlink automatici per DOI e URL + dimensione colonne
    for col in ws.columns:
        col_letter = col[0].column_letter
        max_len = 0
        for cell in col:
            if cell.row == 1 or not cell.value:
                continue
            val = str(cell.value).strip()
            max_len = max(max_len, len(val))

            # Se sembra un DOI (10.x) lo trasformiamo in https://doi.org/...
            if val.startswith("10."):
                cell.hyperlink = f"https://doi.org/{val}"
                cell.font = Font(color="0000EE", underline="single")
            # Se è già un URL completo
            elif val.startswith("http://") or val.startswith("https://"):
                cell.hyperlink = val
                cell.font = Font(color="0000EE", underline="single")

        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    wb.save(excel_path)
    print(f"🎨 Excel formattato e link cliccabili: {excel_path}")


# === Pipeline principale per GitHub, LODCloud, Zotero ===
for csv_path, out_xlsx in files.items():
    print(f"\n📥 Elaboro: {csv_path}")
    df = read_csv_stable(csv_path)

    # Rimuove colonne completamente vuote
    empty_cols = [c for c in df.columns if df[c].isna().all()]
    if empty_cols:
        print(f"[CLEANUP] Rimosse colonne vuote: {empty_cols}")
        df = df.drop(columns=empty_cols)

    # Ordinamento per data e titolo
    date_col = next((c for c in df.columns if any(k in c.lower() for k in ['date', 'time', 'year'])), None)
    name_col = next((c for c in df.columns if any(k in c.lower() for k in ['title', 'name'])), None)
    if date_col:
        try:
            df[date_col] = pd.to_datetime(df[date_col], errors='coerce').dt.tz_localize(None)
            df['Year'] = df[date_col].dt.year
            sort_cols = ['Year', date_col]
            ascending = [False, False]
            if name_col:
                sort_cols.append(name_col)
                ascending.append(True)
            df = df.sort_values(by=sort_cols, ascending=ascending).drop(columns=['Year'])
        except Exception as e:
            print(f"[WARN] Ordinamento non riuscito per {csv_path}: {e}")

    out_dir = Path(out_xlsx).parent
    out_dir.mkdir(parents=True, exist_ok=True)
    df.to_excel(out_xlsx, index=False)
    format_excel_table(out_xlsx)


# === Zenodo completo dal CSV senza rinomina ===
print("\n📚 Elaboro Zenodo direttamente dal CSV completo ...")
df_zen = read_csv_stable(zenodo_csv)

# Pulizia dati
for col in df_zen.columns:
    df_zen[col] = df_zen[col].apply(clean_text)

# Ordinamento per anno e titolo se presenti
year_col = next((c for c in df_zen.columns if 'year' in c.lower()), None)
title_col = next((c for c in df_zen.columns if 'title' in c.lower()), None)
sort_cols = []
ascending = []
if year_col:
    df_zen[year_col] = pd.to_numeric(df_zen[year_col], errors='coerce')
    sort_cols.append(year_col)
    ascending.append(False)
if title_col:
    sort_cols.append(title_col)
    ascending.append(True)
if sort_cols:
    df_zen = df_zen.sort_values(by=sort_cols, ascending=ascending)

# Salva Excel
df_zen.to_excel(zenodo_xlsx, index=False)
format_excel_table(zenodo_xlsx)

print("\n✅ Tutti i file Excel (incluso Zenodo) sono stati creati e formattati correttamente!")
